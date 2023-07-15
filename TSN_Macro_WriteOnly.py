import os
import traceback
from win32com.client import Dispatch
from openpyxl import load_workbook
from datetime import datetime

# define the paths
EXCEL_PATH = 'C:/Users/sayan/Likewize/Operations Data Warehouse - Documents/CUBE_TSN Extract/Transaction Records/Macro_Transactions.xlsx'
MACRO_FILE = 'C:/Users/sayan/Likewize/Operations Data Warehouse - Macro_TsnCube/TSN_Cube.xlsm'
# define the macros
MACRO_NAMES = ['TSN_Master', 'TSN_Shipment', 'TSN_SOH']

def run_macro():
    excel_macro = Dispatch("Excel.Application")
    excel_macro.Visible = False

    for macro in MACRO_NAMES:
        try:
            workbook = excel_macro.Workbooks.Open(Filename=MACRO_FILE, ReadOnly=1)
            excel_macro.Application.Run(macro)
            record_success(macro)

            # Save and Close
            workbook.Save()
            workbook.Close()

        except Exception as e:
            record_failure(macro, traceback.format_exc())

    # Ensure Excel process is closed
    excel_macro.Application.Quit()
    del excel_macro

def record_success(macro_name):
    wb = load_workbook(filename = EXCEL_PATH)
    ws = wb['Sheet1']  # replace 'Sheet1' with the name of the sheet you want to write in
    ws.append([str(datetime.now()), MACRO_FILE, macro_name, "Success", ""])
    wb.save(EXCEL_PATH)

def record_failure(macro_name, error):
    wb = load_workbook(filename = EXCEL_PATH)
    ws = wb['Sheet1']  # replace 'Sheet1' with the name of the sheet you want to write in
    ws.append([str(datetime.now()), MACRO_FILE, macro_name, "Failure", str(error)])
    wb.save(EXCEL_PATH)

run_macro()
