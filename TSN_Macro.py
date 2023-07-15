import os
import traceback
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from win32com.client import Dispatch
from openpyxl import load_workbook
from datetime import datetime
import time

# define the paths
EXCEL_PATH = 'C:/Users/sayan/Likewize/Operations Data Warehouse - Documents/CUBE_TSN Extract/Transaction Records/Macro_Transactions.xlsx'
MACRO_FILE = 'C:/Users/sayan/Likewize/Operations Data Warehouse - Macro_TsnCube/TSN_Cube.xlsm'

# define the macros
MACRO_NAMES = ['TSN_Master', 'TSN_Shipment', 'TSN_SOH']

# define email parameters
FROM_ADDRESS = "supplychainOceania@brightstar.com"
EMAIL_PASSWORD = "W98h4XSNaa"
TO_ADDRESS = ["analyticsandinsights.au@likewize.com"]  # replace with the real recipients
SMTP_SERVER = "smtp.office365.com"
SMTP_PORT = 587

def run_macro():
    excel_macro = Dispatch("Excel.Application")
    excel_macro.Visible = False

    for macro in MACRO_NAMES:
        # Start timing
        start_time = time.time()

        try:
            workbook = excel_macro.Workbooks.Open(Filename=MACRO_FILE, ReadOnly=1)
            excel_macro.Application.Run(macro)

            # Stop timing and calculate elapsed time
            end_time = time.time()
            execution_time = end_time - start_time

            record_success(macro, execution_time)

            # Save and Close
            workbook.Save()
            workbook.Close()

        except Exception as e:
            # Stop timing and calculate elapsed time even in case of an error
            end_time = time.time()
            execution_time = end_time - start_time

            error_message = traceback.format_exc()
            record_failure(macro, error_message, execution_time)
            send_email(macro, error_message)

    # Ensure Excel process is closed
    excel_macro.Application.Quit()
    del excel_macro

def record_success(macro_name, execution_time):
    wb = load_workbook(filename = EXCEL_PATH)
    ws = wb['Sheet1']  # replace 'Sheet1' with the name of the sheet you want to write in
    ws.append([str(datetime.now()), MACRO_FILE, macro_name, "Success", execution_time, ""])
    wb.save(EXCEL_PATH)

def record_failure(macro_name, error, execution_time):
    wb = load_workbook(filename = EXCEL_PATH)
    ws = wb['Sheet1']  # replace 'Sheet1' with the name of the sheet you want to write in
    ws.append([str(datetime.now()), MACRO_FILE, macro_name, "Failure", execution_time, str(error)])
    wb.save(EXCEL_PATH)

def send_email(macro_name, error):
    msg = MIMEMultipart()
    msg['From'] = FROM_ADDRESS
    msg['To'] = ", ".join(TO_ADDRESS)
    msg['Subject'] = "Macro execution failure: " + macro_name

    body = "There was an error during macro execution. \n\n" + error
    msg.attach(MIMEText(body, 'plain'))

    server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
    server.starttls()
    server.login(FROM_ADDRESS, EMAIL_PASSWORD)
    text = msg.as_string()
    server.sendmail(FROM_ADDRESS, TO_ADDRESS, text)
    server.quit()

run_macro()